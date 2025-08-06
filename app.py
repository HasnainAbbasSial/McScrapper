from flask import Flask, render_template, request, jsonify, Response, send_file, session, redirect, url_for
from flask_socketio import SocketIO, emit
import csv
import json
import io
import threading
import time
from datetime import datetime
from scraper import FMCSAScraper
from license_service import license_validator
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')
socketio = SocketIO(app, cors_allowed_origins="*")

# Contact information (easy to modify)
CONTACT_INFO = {
    'email': 'hasnainabbas.contact@gmail.com',
    'whatsapp': '+923070467687'
}

# Global variables to manage scraping state
scraper_instance = None
scraping_active = False
scraped_data = []

# Global variables for license expiry monitoring
license_expired_keys = set()  # Track expired license keys
expiry_monitor_running = False

def is_authenticated():
    if not session.get('authenticated', False):
        return False
    
    # Check if this user's license key has expired
    user_license_key = session.get('license_key', '')
    if user_license_key in license_expired_keys:
        # Clear the session for expired users
        session.clear()
        return False
    
    return True

def get_user_data():
    return {
        'name': session.get('user_name', ''),
        'email': session.get('user_email', ''),
        'license_key': session.get('license_key', '')
    }

def monitor_license_expiry():
    """Background function to check license expiry every minute"""
    global expiry_monitor_running
    expiry_monitor_running = True
    
    while expiry_monitor_running:
        try:
            # Check licenses every 60 seconds (1 minute)
            time.sleep(60)
            
            # Get all current licenses from the system
            licenses = license_validator.get_all_licenses()
            if not licenses:
                continue
            
            current_time = datetime.now()
            newly_expired = []
            
            for license_data in licenses:
                license_key = license_data.get('license_key', '')
                expiry_date_str = license_data.get('expiry_date', '')
                
                if not license_key or not expiry_date_str:
                    continue
                
                try:
                    # Parse expiry date (assuming format: YYYY-MM-DD or DD/MM/YYYY)
                    if '/' in expiry_date_str:
                        expiry_date = datetime.strptime(expiry_date_str, '%d/%m/%Y')
                    else:
                        expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d')
                    
                    # Check if license has expired
                    if current_time > expiry_date:
                        newly_expired.append(license_key)
                        print(f"License expired: {license_key} (expired on {expiry_date_str})")
                        
                        # Send expiry notification to all connected clients with this license
                        socketio.emit('license_expired', {
                            'message': 'Your license has expired. Please contact support to renew.',
                            'license_key': license_key,
                            'expiry_date': expiry_date_str
                        }, room=None)
                        
                except ValueError as e:
                    print(f"Error parsing expiry date for license {license_key}: {e}")
                    continue
            
            # Update the global expired licenses set
            license_expired_keys.update(newly_expired)
            
        except Exception as e:
            print(f"Error in license expiry monitor: {e}")
            time.sleep(30)  # Wait 30 seconds before retrying on error

def start_license_expiry_monitor():
    """Start the background license expiry monitoring thread"""
    global expiry_monitor_running
    if not expiry_monitor_running:
        monitor_thread = threading.Thread(target=monitor_license_expiry, daemon=True)
        monitor_thread.start()
        print("License expiry monitor started")

@app.route('/')
def index():
    if not is_authenticated():
        return redirect(url_for('login_page'))
    
    user_data = get_user_data()
    return render_template('index.html', user=user_data)

@app.route('/login')
def login_page():
    if is_authenticated():
        return redirect(url_for('index'))
    return render_template('login.html', contact=CONTACT_INFO)

@app.route('/validate_license', methods=['POST'])
def validate_license():
    data = request.get_json()
    license_key = data.get('license_key', '').strip()
    
    if not license_key:
        return jsonify({'success': False, 'message': 'Please enter a license key.'})
    
    # Store license key temporarily for the next step
    session['temp_license_key'] = license_key
    
    return jsonify({
        'success': True, 
        'message': 'License key verified. Please enter your email.',
        'next_step': 'email_verification'
    })

@app.route('/validate_email', methods=['POST'])
def validate_email():
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    license_key = session.get('temp_license_key', '')
    
    if not email:
        return jsonify({'success': False, 'message': 'Please enter your email address.'})
    
    if not license_key:
        return jsonify({'success': False, 'message': 'License key not found. Please start over.'})
    
    # Validate license key and email combination
    result = license_validator.validate_license(license_key, email)
    
    if result['valid']:
        # Store user session data with unique session ID
        import uuid
        session_id = str(uuid.uuid4())
        session['authenticated'] = True
        session['user_name'] = result.get('name', 'User')
        session['user_email'] = email
        session['license_key'] = license_key
        session['session_id'] = session_id
        session.pop('temp_license_key', None)  # Remove temporary key
        
        # Start license expiry monitor if not already running
        start_license_expiry_monitor()
        
        return jsonify({
            'success': True, 
            'message': f'Welcome {result.get("name", "User")}! Access granted.',
            'redirect_url': url_for('index')
        })
    else:
        return jsonify({'success': False, 'message': result['message']})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@socketio.on('start_scraping')
def handle_start_scraping(data):
    # Check if user is authenticated
    if not is_authenticated():
        emit('error', {'message': 'Access denied. Please login with your license key.'})
        return
    global scraper_instance, scraping_active, scraped_data
    
    if scraping_active:
        emit('error', {'message': 'Scraping is already in progress'})
        return
    
    try:
        start_mc = int(data['start_mc'])
        end_mc = int(data.get('end_mc', 0)) if data.get('end_mc') else None
        entity_type = data['entity_type']
        
        # Reset scraped data
        scraped_data = []
        scraping_active = True
        
        # Create scraper instance
        scraper_instance = FMCSAScraper(start_mc, end_mc, entity_type)
        
        # Start scraping in a separate thread
        thread = threading.Thread(target=run_scraping)
        thread.daemon = True
        thread.start()
        
        emit('scraping_started', {'message': 'Scraping started successfully'})
        
    except ValueError:
        emit('error', {'message': 'Invalid MC number format'})
    except Exception as e:
        emit('error', {'message': f'Error starting scraping: {str(e)}'})

@socketio.on('stop_scraping')
def handle_stop_scraping():
    global scraping_active, scraper_instance
    
    scraping_active = False
    if scraper_instance:
        scraper_instance.stop()
    
    emit('scraping_stopped', {'message': 'Scraping stopped'})

def run_scraping():
    global scraper_instance, scraping_active, scraped_data
    
    def on_progress(current_mc, status, data=None):
        socketio.emit('progress_update', {
            'current_mc': current_mc,
            'status': status,
            'data': data
        })
        
        if data and status == 'valid':
            scraped_data.append(data)
            socketio.emit('data_update', {'data': data, 'total_count': len(scraped_data)})
    
    def on_complete():
        global scraping_active
        scraping_active = False
        socketio.emit('scraping_complete', {'total_found': len(scraped_data)})
    
    try:
        scraper_instance.scrape(on_progress, on_complete)
    except Exception as e:
        socketio.emit('error', {'message': f'Scraping error: {str(e)}'})
        scraping_active = False

@app.route('/export/<format>')
def export_data(format):
    if not is_authenticated():
        return jsonify({'error': 'Access denied. Please login with your license key.'}), 403
    global scraped_data
    
    if not scraped_data:
        return jsonify({'error': 'No data to export'}), 400
    
    if format == 'csv':
        return export_csv()
    elif format == 'xlsx':
        return export_xlsx()
    elif format == 'txt':
        return export_txt()
    else:
        return jsonify({'error': 'Invalid format'}), 400

def export_csv():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        'mc_number', 'usdot_number', 'legal_name', 'physical_address', 
        'phone_number', 'email'
    ])
    writer.writeheader()
    for row in scraped_data:
        writer.writerow(row)
    
    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=fmcsa_data.csv'}
    )
    return response

def export_xlsx():
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FMCSA Data"
        
        # Headers
        headers = ['MC Number', 'USDOT Number', 'Legal Name', 'Physical Address', 'Phone Number', 'Email']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row_num, row_data in enumerate(scraped_data, 2):
            ws.cell(row=row_num, column=1, value=row_data.get('mc_number', ''))
            ws.cell(row=row_num, column=2, value=row_data.get('usdot_number', ''))
            ws.cell(row=row_num, column=3, value=row_data.get('legal_name', ''))
            ws.cell(row=row_num, column=4, value=row_data.get('physical_address', ''))
            ws.cell(row=row_num, column=5, value=row_data.get('phone_number', ''))
            ws.cell(row=row_num, column=6, value=row_data.get('email', ''))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=fmcsa_data.xlsx'}
        )
        return response
        
    except ImportError:
        return jsonify({'error': 'openpyxl not available for Excel export'}), 500
    except Exception as e:
        return jsonify({'error': f'Excel export failed: {str(e)}'}), 500

def export_txt():
    output = io.StringIO()
    output.write("FMCSA Data Export\n")
    output.write("=" * 50 + "\n\n")
    
    for i, row in enumerate(scraped_data, 1):
        output.write(f"Record {i}:\n")
        output.write(f"MC Number: {row.get('mc_number', 'N/A')}\n")
        output.write(f"USDOT Number: {row.get('usdot_number', 'N/A')}\n")
        output.write(f"Legal Name: {row.get('legal_name', 'N/A')}\n")
        output.write(f"Physical Address: {row.get('physical_address', 'N/A')}\n")
        output.write(f"Phone Number: {row.get('phone_number', 'N/A')}\n")
        output.write(f"Email: {row.get('email', 'N/A')}\n")
        output.write("-" * 30 + "\n\n")
    
    response = Response(
        output.getvalue(),
        mimetype='text/plain',
        headers={'Content-Disposition': 'attachment; filename=fmcsa_data.txt'}
    )
    return response

@app.route('/status')
def get_status():
    return jsonify({
        'scraping_active': scraping_active,
        'data_count': len(scraped_data)
    })

@app.route('/health')
def health_check():
    return jsonify({
        'status': 'healthy',
        'message': 'FMCSA Scraper is running'
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    socketio.run(app, host='0.0.0.0', port=port, debug=False)
